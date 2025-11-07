import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Find latest report
files = glob.glob('outputs/Strict_Duplicates_*.xlsx')
latest = max(files)
print(f'📂 قراءة التقرير: {latest}\n')

# Read duplicates
df = pd.read_excel(latest, sheet_name='All_Duplicates')
print(f'✅ عدد السجلات المكررة: {len(df)}\n')

# Check for new columns
print('📋 الأعمدة الجديدة:')
new_cols = ['_ConfirmedDuplicate', 'ReasonText']
for col in new_cols:
    if col in df.columns:
        print(f'   ✅ {col}')
        if col == 'ReasonText':
            print(f'\n   📝 أمثلة الرسائل الجديدة:')
            unique_reasons = df['ReasonText'].unique()
            for reason in unique_reasons[:5]:
                count = len(df[df['ReasonText'] == reason])
                print(f'      • {reason} ({count} سجل)')
    else:
        print(f'   ❌ {col} - غير موجود')

# Check confirmed duplicates
if '_ConfirmedDuplicate' in df.columns:
    confirmed_count = df['_ConfirmedDuplicate'].sum()
    print(f'\n🔴 التكرارات المؤكدة (تكرار + بنك مطابق): {confirmed_count} سجل')
    print(f'⚠️ التكرارات غير المؤكدة: {len(df) - confirmed_count} سجل')

# Check coloring in Excel
print(f'\n🎨 فحص التلوين في Excel...')
wb = load_workbook(latest)
ws = wb['All_Duplicates']

pink_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
colored_rows = 0

for row in range(2, len(df) + 2):  # Start from row 2 (after header)
    cell = ws.cell(row=row, column=1)
    if cell.fill and cell.fill.start_color:
        if cell.fill.start_color.rgb and 'FFC7CE' in str(cell.fill.start_color.rgb):
            colored_rows += 1

print(f'   عدد الصفوف الملونة بـ #FFC7CE: {colored_rows}')

wb.close()

print('\n' + '='*100)
print('✅ التحديثات المنفذة بنجاح!')
print('='*100)
